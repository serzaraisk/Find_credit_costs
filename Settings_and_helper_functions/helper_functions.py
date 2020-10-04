import os
import pandas as pd
import re
import math
from itertools import combinations
from openpyxl import load_workbook
from Settings_and_helper_functions import Cell_formating, Print_sub_functions, Settings
from Stars import Stars_settings


def get_credit_accounts_list(bank, INN, debug_lvl):
    """ Возвращает список кредитных счетов из Stars по указанному банку и ИНН заемщика

    :param bank: Банк, из БД которой будет выгружена выписка счетов по клиенту
    :param INN: ИНН заемщика
    :param debug_lvl: нужен для Print_sub_function, onределяет работу print
    :return: список кредитных счетов заемщика
    """
    accounts_list = Stars_settings.sql_get_accounts_based_on_INN(bank, INN, debug_lvl)['Счёт']
    # фильтр по счетам (441** – 457**) и убираются счета, у которых счет второго порядка 15
    filter_credit_accounts = (accounts_list.str.contains(r'(^44[1-9])|(^45[0-7]).*')) & \
                             (~ accounts_list.str.contains(r'(^\d\d\d15\d*)'))
    Print_sub_functions.print_account_list_kicked(accounts_list, filter_credit_accounts, debug_lvl)
    return accounts_list[filter_credit_accounts].unique()


def get_excel_file(path=""):
    """Ищет Excel файл в указанной папке (если не указан путь, то в папке, где находится скрипт)

    :param path: (optional) путь к исходному файлу (Перечень ссудных счетов второго порядка)
    :return: имя файла или текст с ошибкой

    """
    path = os.getcwd() if path == "" else path
    list_of_cwd = os.listdir(path)
    for file in list_of_cwd:
        if ".xls" in file and file != 'output_1.xlsx':
            return file if path != os.getcwd() else path + "\\" + file
    return "В выбранной папке нет xls файла"


def parse_excel_file_to_pandas(input_file, worksheet=0):
    """Переводит excel файл в dataframe

    :param input_file: excel файл, который был получен путем выполнения функции get_excel_file
    :param worksheet: название листа, с которого забирается таблица
    :return: dataframe c данными excel
    """
    if worksheet != 0:
        input_file = load_workbook(input_file)[worksheet].values
    else:
        input_file = load_workbook(input_file).worksheets[worksheet].values
    columns = next(input_file)
    output_df = pd.DataFrame(input_file, columns=columns)
    index_dict = {account: number + 1 for number, account in enumerate(output_df['Номер исследуемого счёта'].unique())}
    output_df['Порядковый номер кредитного договора'] = output_df['Номер исследуемого счёта'].map(index_dict)
    return output_df


def create_row(cost, row_to_add):
    """Заполняет первоначальными ключами и значениями dict с информацией о расходах

    :param cost: строка из выписки расчетного счета(фактически, строка расходов)
    :param row_to_add: dict, в который заполняются данные по расходованию кредитных средств
    :return: dict со сформированными ключами и значениями
    """
    no_creation = 'Определить направление расходования не представляется возможным. Необходимо посмотреть вручную'
    if cost[1].get('Направление (Целевое назначение)', -1) != no_creation:
        row_to_add['Направление (Дата перечисления)'] = [str(cost[1]['Дата операции'])]
        row_to_add['Направление (Сумма перечисления)'] = [str(cost[1]['Эквивалент суммы по дебету'])]
        row_to_add['Направление(Название)'] = [cost[1]['Наименование внутреннего корреспондирующего счёта']]
        row_to_add['Направление(ИНН)'] = [str(cost[1]['ИНН контрагента'])]
        row_to_add['Направление(Банк)'] = [str(cost[1]['Наименование банка контрагента'])]
        # временное решение: если счет начинается на 30, то это межбанк
        if str(cost[1]['Номер внутреннего корреспондирующего счёта']).startswith('30'):
            row_to_add['Направление(Номер корреспондирующего счёта)'] = [str(
                cost[1]['Номер счёта контрагента'])]
        else:
            row_to_add['Направление(Номер корреспондирующего счёта)'] = [str(
                cost[1]['Номер внутреннего корреспондирующего счёта'])]
        row_to_add['Направление(Назначение платежа)'] = [cost[1]['Назначение платежа']]
        goal_purpose = identify_account(cost[1])
        row_to_add['Направление (Целевое назначение)'] = [goal_purpose]
    else:
        create_zero_row(row_to_add)


def add_row(cost, row_to_add):
    """Дополняет dict с информацией о расходах информацией из дополнительных строк расходов

    :param cost: строка из выписки расчетного счета(фактически, строка расходов)
    :param row_to_add: dict, в который заполняются данные по расходованию кредитных средств
    :return: dict со сформированными ключами и значениями
    """
    row_to_add['Направление (Дата перечисления)'] = row_to_add['Направление (Дата перечисления)'] + [
        cost[1]['Дата операции']]
    row_to_add['Направление (Сумма перечисления)'] = row_to_add['Направление (Сумма перечисления)'] \
                                                     + [cost[1]['Эквивалент суммы по дебету']]
    # print('прошло 2 добавление')
    row_to_add['Направление(Название)'] = row_to_add['Направление(Название)'] + \
                                              [cost[1]['Наименование внутреннего корреспондирующего счёта']]
    # print('прошло 3 добавление')
    row_to_add['Направление(ИНН)'] = row_to_add['Направление(ИНН)'] + [cost[1]['ИНН контрагента']]
    # print('прошло 4 добавление')
    row_to_add['Направление(Банк)'] = row_to_add['Направление(Банк)'] + [cost[1]['Наименование банка контрагента']]
    # print('прошло 5 добавление')
    if str(cost[1]['Номер внутреннего корреспондирующего счёта']).startswith('30'):
        row_to_add['Направление(Номер корреспондирующего счёта)'] = row_to_add[
                                                                        'Направление(Номер корреспондирующего счёта)'] \
                                                                    + [cost[1]['Номер счёта контрагента']]
    else:
        row_to_add['Направление(Номер корреспондирующего счёта)'] = row_to_add[
                                                                        'Направление(Номер корреспондирующего счёта)'] \
                                                                    + [cost[1][
                                                                           'Номер внутреннего корреспондирующего счёта']]
    row_to_add['Направление(Назначение платежа)'] = row_to_add['Направление(Назначение платежа)'] + [
        cost[1]['Назначение платежа']]
    # print('прошло 7 добавление')
    goal_purpose = identify_account(cost[1])
    row_to_add['Направление (Целевое назначение)'] = row_to_add['Направление (Целевое назначение)'] + [goal_purpose]
    row_to_add['Направление расходования ссудной задолженности'] = row_to_add[
                                                                       'Направление расходования ссудной задолженности'] \
                                                                   + [row_to_add[
                                                                          'Направление расходования ссудной задолженности'][
                                                                          0]]


def create_zero_row(row_to_add):
    """Создает пустую строку: необходимо, если ссуда уходит на межбанк или деньги выдаются в кассе"""
    row_to_add['Направление (Дата перечисления)'] = ['Не применимо']
    row_to_add['Направление (Сумма перечисления)'] = ['Не применимо']
    row_to_add['Направление(Название)'] = ['Не применимо']
    row_to_add['Направление(ИНН)'] = ['Не применимо']
    row_to_add['Направление(Банк)'] = ['Не применимо']
    row_to_add['Направление(Назначение платежа)'] = ['Не применимо']
    row_to_add['Направление (Целевое назначение)'] = ['Не применимо']
    row_to_add['Направление(Номер корреспондирующего счёта)'] = ['Не применимо']


def create_inner_end_row(cost, row_to_add):
    """Создает строку для конечного получателя средств"""
    row_to_add['Конечный получатель (Название)'] = [cost['Наименование контрагента']]
    row_to_add['Конечный получатель (ИНН)'] = [str(cost['ИНН контрагента'])]
    row_to_add['Конечный получатель (Дата перечисления контрагенту)'] = [str(cost['Дата операции'])]
    row_to_add['Конечный получатель (Банк)'] = [str(cost['Наименование банка контрагента'])]
    row_to_add['Конечный получатель (Сумма перечисления)'] = [str(cost['Эквивалент суммы по дебету'])]
    row_to_add['Конечный получатель (Назначение платежа)'] = [cost['Назначение платежа']]
    row_to_add['Конечный получатель (Номер счета контрагента)'] = cost['Номер внутреннего корреспондирующего счёта']
    if str(cost['Номер внутреннего корреспондирующего счёта']).startswith('30'):
        try:
            row_to_add['Конечный получатель (Номер счета контрагента)'] = [str(cost['Номер счёта контрагента'])]
        except KeyError:
            row_to_add['Конечный получатель (Номер счета контрагента)'] = ['Не применимо']
    else:
        row_to_add['Конечный получатель (Номер счета контрагента)'] = [
            str(cost['Номер внутреннего корреспондирующего счёта'])]
    row_to_add['Конечный получатель (Целевое назначение)'] = [identify_account(cost)]
    if row_to_add['Конечный получатель (Целевое назначение)'] == ['Рефинансирование ссудной задолженности']:
        row_to_add['Конечный получатель (Название)'] = 'Банк'
    row_to_add['Комментарий'] = ['']


def create_inner_end_zero_row():
    """Создает пустую конечную строку получателя денежных средств(в случае если по расчетному счету нет выписки)"""
    zero_row = {'Конечный получатель (Название)': ['Не применимо'],
                'Конечный получатель (ИНН)': ['Не применимо'],
                'Конечный получатель (Дата перечисления контрагенту)': ['Не применимо'],
                'Конечный получатель (Банк)': ['Не применимо'],
                'Конечный получатель (Сумма перечисления)': ['Не применимо'],
                'Конечный получатель (Назначение платежа)': ['Не применимо'],
                'Конечный получатель (Номер счета контрагента)': ['Не применимо'],
                'Конечный получатель (Целевое назначение)': ['Не применимо'], 'Комментарий': ['Не применимо']}
    return pd.DataFrame(zero_row)


def identify_account(cost):
    """Определяет сценарий расходования средств на основе кор счета для расчетного счета контрагента"""

    regex_credit_accounts = re.compile(r'^((?:44[1-9])|(?:45[0-8])).*$')
    account = cost['Номер внутреннего корреспондирующего счёта']
    if str(cost['Номер внутреннего корреспондирующего счёта']).startswith('202'):
        goal_purpose = 'Через кассу Банка'
    elif str(cost['Номер внутреннего корреспондирующего счёта']).startswith('30102'):
        goal_purpose = 'Кредитные средства перечислены на расчетный / текущий счет заемщика или контрагента в другую КО'
    elif str(cost['Номер внутреннего корреспондирующего счёта']).startswith('40') or \
            str(cost['Номер внутреннего корреспондирующего счёта']).startswith('303'):
        goal_purpose = 'Кредитные средства перечислены на расчетный / текущий счет контрагента в Банке'
    elif re.search(regex_credit_accounts, cost['Номер внутреннего корреспондирующего счёта']) is not None:
        goal_purpose = 'Рефинансирование ссудной задолженности'
    elif str(cost['Номер внутреннего корреспондирующего счёта']).startswith('524'):
        goal_purpose = 'Кредитные средства направлены на приобретение ценных бумаг Банка'
    else:
        goal_purpose = 'Направление расходования кредитных средств необходимо посмотреть вручную'
    return goal_purpose


def parse_act_date_contract(payment_desc):
    """Ищет дату и номер договора в описании"""
    act_regex = re.compile(r'(?:по кредитному договору|по договору кредитной линии| по кредитной линии)(№?N?.*)')
    data_regex = re.compile(r'от (\d\d.\d\d.\d\d\d?\d?)')
    act = re.search(act_regex, payment_desc)
    act_without_date_regex = re.compile(r'(.*)(от .*)')
    try:
        act_without_date = re.search(act_without_date_regex, act.group(1))
        act = act_without_date.group(1)
    except (AttributeError, TypeError):
        act = payment_desc
    try:
        date = re.search(data_regex, payment_desc)
        date = date.group(1)
    except (AttributeError, TypeError):
        date = ''
    return act, date


def load_to_excel(output_table, accounts_dict, debug_lvl):
    """Обрабатывает output_table для загрузки данных в excel"""
    output_table = Settings.reorder_columns(output_table)
    """Создает excel файл с итоговой таблицей"""
    writer = pd.ExcelWriter('output_1.xlsx', engine='xlsxwriter', date_format='YYYY-MM-DD’')
    # Преобразует часть столбцов, чтобы на них можно было применить форматирование
    output_table['Дата выдачи транша'] = output_table['Дата выдачи транша'].astype('str')
    output_table['Направление (Дата перечисления)'] = output_table['Направление (Дата перечисления)'].astype('str')

    output_table['Направление (Сумма перечисления)'] = output_table['Направление (Сумма перечисления)'].astype('float',
                                                                                                               errors='ignore')

    # Переименование части столбцов
    output_table.rename({'Направление(Название)': 'Направление (Наименование получателя средств)',
                         'Направление(ИНН)': 'Направление (ИНН получателя средств)',
                         'Направление(Банк)': 'Направление (Банк получателя средств)',
                         'Направление(Номер корреспондирующего счёта)': 'Направление (Номер корреспондирующего счёта)',
                         'Направление(Назначение платежа)': 'Направление (Назначение платежа)',
                         'Комментарий': 'Цепочка получателей средств (% от предыдущего транжа)'}, inplace=True, axis=1)
    # Convert the dataframe to an XlsxWriter Excel object.
    output_table.to_excel(writer, sheet_name='Sheet1')
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    # Форматирование таблицы
    worksheet.add_table(1, 0, output_table.shape[0], output_table.shape[1], {'header_row': True,
                                                                             'style': 'Table Style Light 8'})
    # Форматирование загаловка(создание форматов и присвоение форматов)
    cell_format_header_blue = Cell_formating.cell_format_for_header_blue(workbook)
    cell_format_header_medium_blue = Cell_formating.cell_format_for_header_medium_blue(workbook)
    cell_format_header_purple = Cell_formating.cell_format_for_header_purple(workbook)
    cell_format_header_green = Cell_formating.cell_format_for_header_green(workbook)
    cell_format_second_row = Cell_formating.cell_format_for_second_row(workbook)
    dict_header = {'Порядковый № исследуемого ИНН': cell_format_header_medium_blue,
                   'Порядковый номер исследуемого ссудного счета': cell_format_header_medium_blue,
                   "Порядковый № кредитного транжа": cell_format_header_medium_blue,
                   'Порядковый номер первого (после заемщика) получателя средств': cell_format_header_medium_blue,
                   'Порядковый номер конечного получателя средств': cell_format_header_medium_blue,
                   "Наименование заемщика": cell_format_header_blue,
                   'ИНН заемщика': cell_format_header_blue,
                   'Номер кредитного договора': cell_format_header_blue,
                   'Дата кредитного договора': cell_format_header_blue,
                   'Валюта кредита': cell_format_header_blue,
                   'Сумма выданных в исследуемом периоде траншей': cell_format_header_blue,
                   'Дата выдачи транша': cell_format_header_blue,
                   'Сумма транша': cell_format_header_blue,
                   'Номер внутреннего корреспондирующего счёта': cell_format_header_blue,
                   'Направление расходования ссудной задолженности': cell_format_header_purple,
                   'Направление (Дата перечисления)': cell_format_header_purple,
                   'Направление (Сумма перечисления)': cell_format_header_purple,
                   'Направление (Наименование получателя средств)': cell_format_header_purple,
                   'Направление (ИНН получателя средств)': cell_format_header_purple,
                   'Направление (Банк получателя средств)': cell_format_header_purple,
                   'Направление (Номер корреспондирующего счёта)': cell_format_header_purple,
                   'Направление (Назначение платежа)': cell_format_header_purple,
                   'Направление (Целевое назначение)': cell_format_header_purple,
                   'Конечный получатель (Название)': cell_format_header_green,
                   'Конечный получатель (ИНН)': cell_format_header_green,
                   'Конечный получатель (Дата перечисления контрагенту)': cell_format_header_green,
                   'Конечный получатель (Банк)': cell_format_header_green,
                   'Конечный получатель (Сумма перечисления)': cell_format_header_green,
                   'Конечный получатель (Назначение платежа)': cell_format_header_green,
                   'Конечный получатель (Номер счета контрагента)': cell_format_header_green,
                   'Конечный получатель (Целевое назначение)': cell_format_header_green,
                   'Цепочка получателей средств (% от предыдущего транжа)': cell_format_header_green,
                   }
    # переписывание название колонок(иначе не работает фильтры в excel)
    for num, cell in enumerate(output_table.columns.values):
        cell_format_header = dict_header[cell]
        worksheet.write_string(0, num + 1, cell, cell_format_header)
        worksheet.write_string(1, num + 1, str(num + 1), cell_format_second_row)

    # Формитирование колонок
    cell_format = Cell_formating.cell_format_main(workbook)
    cell_format_numeric = Cell_formating.cell_format_for_numeric(workbook)
    cell_format_for_date = Cell_formating.cell_format_for_date(workbook)

    worksheet.set_column('B:B', 14.44, cell_format)  # Порядковый № исследуемого ИНН
    worksheet.set_column('C:C', 21, cell_format)  # Порядковый номер исследуемого ссудного счета
    worksheet.set_column('D:D', 13.89, cell_format)  # Порядковый № транжа
    worksheet.set_column('E:E', 18.44, cell_format)  # Порядковый номер первого (после заемщика) получателя средств
    worksheet.set_column('F:F', 18, cell_format)  # Порядковый номер конечного получателя средств
    worksheet.set_column('G:G', 16.22, cell_format)  # Наименование заемщика
    worksheet.set_column('H:H', 10.56, cell_format)  # ИНН заемщика
    worksheet.set_column('I:I', 22, cell_format)  # Номер кредитного договора
    worksheet.set_column('J:J', 15.22, cell_format_for_date)  # Дата кредитного договора
    worksheet.set_column('K:K', 8.22, cell_format)  # Валюта кредита
    worksheet.set_column('L:L', 22.33, cell_format_numeric)  # Сумма выданных в исследуемом периоде траншей
    worksheet.set_column('M:M', 19.33, cell_format)  # Дата выдачи транша
    worksheet.set_column('N:N', 13.33, cell_format_numeric)  # Сумма транша
    worksheet.set_column('O:O', 25.22, cell_format)  # Номер внутреннего корреспондирующего счёта
    worksheet.set_column('P:P', 48.33, cell_format)  # Направление расходования ссудной задолженности
    worksheet.set_column('Q:Q', 18.67, cell_format)  # Направление (Дата перечисления)
    worksheet.set_column('R:R', 23.33, cell_format_numeric)  # Направление (Сумма перечисления)
    worksheet.set_column('S:S', 22.67, cell_format)  # Направление(Название)
    worksheet.set_column('T:T', 16.11, cell_format)  # Направление(ИНН)
    worksheet.set_column('U:U', 19.89, cell_format)  # Направление(Банк)
    worksheet.set_column('V:V', 26.89, cell_format)  # Направление(Номер корреспондирующего счёта)
    worksheet.set_column('W:W', 33.89, cell_format)  # Направление(Назначение платежа)
    worksheet.set_column('X:X', 33.22, cell_format)  # Направление (Целевое назначение)
    worksheet.set_column('Y:Y', 23, cell_format)  # Конечный получатель (Название)
    worksheet.set_column('Z:Z', 18.44, cell_format)  # Конечный получатель (ИНН)
    worksheet.set_column('AA:AA', 29.56, cell_format)  # Конечный получатель (Банк)
    worksheet.set_column('AB:AB', 24.22, cell_format)  # Конечный получатель (Банк)
    worksheet.set_column('AC:AC', 31, cell_format_numeric)  # Конечный получатель (Сумма перечисления)
    worksheet.set_column('AD:AD', 24.89, cell_format)  # Конечный получатель (Назначение платежа)
    worksheet.set_column('AE:AE', 25.56, cell_format)  # Конечный получатель (Номер счета контрагента)
    worksheet.set_column('AF:AF', 25.56, cell_format)  # Комментарий
    worksheet.set_column('AG:AG', 41.78, cell_format)  # Комментарий

    # Дополнительные настройки
    worksheet.set_column('A:A', 0)
    worksheet.freeze_panes(1, 6)
    worksheet.set_zoom(80)

    # загрузка выписок по анализируемым счетам в случае debug_lvl > 2
    if debug_lvl > 2:
        for account, account_df in accounts_dict.items():
            if 'clear' in account:
                account = account[:-5]
                account_df['Дата операции'] = account_df['Дата операции'].astype('str')
                account_df.to_excel(writer, sheet_name=account)
                worksheet = writer.sheets[account]
                worksheet.add_table(0, 0, account_df.shape[0], account_df.shape[1], {'header_row': False,
                                                                                     'style': 'Table Style Light 8'})
                for num, cell in enumerate(account_df.columns.values):
                    cell_format_header = cell_format_header_blue
                    worksheet.write_string(0, num + 1, cell, cell_format_header)
                worksheet.set_column('A:A', 0)
                worksheet.set_column('B:B', 30.11, cell_format)  # Номер исследуемого счёта
                worksheet.set_column('C:C', 16.22, cell_format)  # Дата операции
                worksheet.set_column('D:D', 12.22, cell_format)  # Время операции
                worksheet.set_column('E:E', 16.56, cell_format)  # Наименование исследуемого клиента
                worksheet.set_column('F:F', 12.33, cell_format)  # ИНН исследуемого клиента
                worksheet.set_column('G:G', 21.22, cell_format)  # Наименование внутреннего корреспондирующего клиента
                worksheet.set_column('H:H', 21.56, cell_format)  # ИНН внутреннего корреспондирующего клиента
                worksheet.set_column('I:I', 20.67, cell_format)  # Наименование внутреннего корреспондирующего счёта
                worksheet.set_column('J:J', 25, cell_format)  # Номер внутреннего корреспондирующего счёта
                worksheet.set_column('K:K', 17, cell_format)  # Наименование контрагента
                worksheet.set_column('L:L', 17.56, cell_format)  # ИНН контрагента
                worksheet.set_column('M:M', 15.56, cell_format)  # Тип счёта контрагента
                worksheet.set_column('N:N', 21.22, cell_format)  # Номер счёта контрагента
                worksheet.set_column('O:O', 14.56, cell_format)  # Наименование банка контрагента
                worksheet.set_column('P:P', 13.44, cell_format)  # ИНН банка контрагента
                worksheet.set_column('Q:Q', 13, cell_format)  # БИК банка контрагента
                worksheet.set_column('R:R', 18.44, cell_format)  # Корреспондентский счёт банка контрагента
                worksheet.set_column('S:S', 15, cell_format)  # Номер документа операции
                worksheet.set_column('T:T', 9.67, cell_format)  # Год операции
                worksheet.set_column('U:U', 12.11, cell_format)  # Месяц операции
                worksheet.set_column('V:V', 11.67, cell_format)  # День операцииС
                worksheet.set_column('W:W', 15, cell_format_numeric)  # Сумма в валюте счёта по дебету
                worksheet.set_column('X:X', 14, cell_format_numeric)  # Сумма в валюте счёта по кредиту
                worksheet.set_column('Y:Y', 12, cell_format_numeric)  # Входящий остаток
                worksheet.set_column('Z:Z', 12.22, cell_format_numeric)  # Эквивалент суммы по дебету
                worksheet.set_column('AA:AA', 12.56, cell_format_numeric)  # Эквивалент суммы по кредиту
                worksheet.set_column('AB:AB', 11.78, cell_format_numeric)  # Исходящий остаток
                worksheet.set_column('AC:AC', 10.22, cell_format)  # Курс обмена валюты
                worksheet.set_column('AD:AD', 88, cell_format)  # Назначение платежа
                worksheet.set_column('AE:AE', 16.22, cell_format)  # Идентификатор операции
                worksheet.set_column('AF:AF', 22.22, cell_format)  # Порядковый номер кредитного договора
                worksheet.set_zoom(80)
    writer.save()


def format_percentages(x):
    """Переводит проценты из float в string"""
    if x['percentage'] != '':
        return "{0:.0%}".format(x['percentage'])
    else:
        return ''

def filter_used_accounts_based_on_id(initial_df, used_df):
    """Отфильтровывает использованные для конечной таблицы строчки банковской выписки из первоначальной банковской выписки"""
    return initial_df[~initial_df['Идентификатор операции'].isin(used_df['Идентификатор операции'])]


def knapsack_filter_sum_function(df, cash_flow):
    """Функция возвращает таблицу с суммой по дебету наиболее приближенной к сумме анализируемого транжа

    :param df: анализируемый текущий счет
    :param cash_flow: данные о анализируемом транже
    :return: отфильтрованную df, на основе алгоритма knapsack
    """
    def knapsack(W, wt, val, n):
        """Ищет максимально приближенную сумму отдельных элементов последовательности к заданной сумме"""
        # Base Case
        if n == 0 or W == 0 :
            return 0

        # If weight of the nth item is more than Knapsack of capacity
        # W, then this item cannot be included in the optimal solution
        if wt[n-1] > W:
            return knapsack(W, wt, val, n-1)

        # return the maximum of two cases:
        # (1) nth item included
        # (2) not included
        else:
            return max(val[n-1] + knapsack(W-wt[n-1], wt, val, n-1),
                       knapsack(W, wt, val, n-1))

    def get_knapsack_values(sum_value_, values_list,length_of_values):
        """Возвращает список элементов из которых может получиться сумма из knapsack алгоритма"""
        for number_of_elements in range(length_of_values+1):
            for elements in list(combinations(values_list, number_of_elements)):
                if sum(elements) == sum_value_:
                    return elements

    value_list = list(df['Эквивалент суммы по дебету'])
    weight_list = list(df['Эквивалент суммы по дебету'])
    try:
        weight_max = int(math.ceil(float(cash_flow['Сумма транша'])))
    except:
        weight_max = int(math.ceil(float(cash_flow['Эквивалент суммы по дебету'])))
    length_of_value_list = len(value_list)
    sum_value = knapsack(weight_max, weight_list, value_list, length_of_value_list)
    sum_list = get_knapsack_values(sum_value, value_list, length_of_value_list)

    return df.iloc[[value_list.index(value) for value in sum_list]]


def remove_min_element_function(output_table,cash_flow):
    """Откидывает расходования пока сумма не будет давать 130% от транжа или пока не останется один элемент"""
    n = output_table.shape[0]
    try:
        while True:
            if output_table['Эквивалент суммы по дебету'].sum() <= 1.3 * cash_flow['Сумма транша']:
                return output_table
            else:
                n -= 1
                output_table = output_table.nlargest(n, 'Эквивалент суммы по дебету', keep='first')
            if n == 1:
                return output_table
    except KeyError:
        while True:
            if output_table['Эквивалент суммы по дебету'].sum() <= 1.3 * cash_flow['Эквивалент суммы по дебету']:
                return output_table
            else:
                n -= 1
                output_table = output_table.nlargest(n, 'Эквивалент суммы по дебету', keep='first')
            if n == 1:
                return output_table
